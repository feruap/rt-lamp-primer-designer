
"""
Dialog Windows

Various dialog windows for the RT-LAMP GUI application.
"""

import os
from pathlib import Path
from typing import List, Dict, Any

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTextEdit,
    QGroupBox, QFormLayout, QSpinBox, QDoubleSpinBox, QCheckBox, QComboBox,
    QLineEdit, QFileDialog, QMessageBox, QTabWidget, QWidget, QProgressBar,
    QDialogButtonBox, QRadioButton, QButtonGroup
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer
from PySide6.QtGui import QFont, QPixmap, QIcon

from rt_lamp_app.design.primer_design import LampPrimerSet
from rt_lamp_app.logger import get_logger


class AboutDialog(QDialog):
    """About dialog for the application."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("About RT-LAMP Primer Designer")
        self.setFixedSize(500, 400)
        
        layout = QVBoxLayout(self)
        
        # Application info
        info_layout = QVBoxLayout()
        
        title_label = QLabel("RT-LAMP Primer Designer")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        info_layout.addWidget(title_label)
        
        version_label = QLabel("Version 1.0.0")
        version_label.setAlignment(Qt.AlignCenter)
        info_layout.addWidget(version_label)
        
        description = QTextEdit()
        description.setReadOnly(True)
        description.setMaximumHeight(200)
        description.setPlainText("""
RT-LAMP Primer Designer is a comprehensive tool for designing 
reverse transcription loop-mediated isothermal amplification 
(RT-LAMP) primers.

Features:
• Automated primer design for F3, B3, FIP, BIP primers
• Optional loop primer design (LF, LB)
• Thermodynamic analysis and optimization
• Specificity checking and cross-reactivity analysis
• Geometric constraint validation
• Export capabilities for laboratory use

Developed by the RT-LAMP Team
Licensed under MIT License
        """)
        info_layout.addWidget(description)
        
        layout.addLayout(info_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)
        button_layout.addWidget(close_button)
        
        layout.addLayout(button_layout)


class SettingsDialog(QDialog):
    """Settings dialog for application preferences."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Settings")
        self.setMinimumSize(600, 500)
        
        self.setup_ui()
        self.load_current_settings()
    
    def setup_ui(self):
        """Setup the settings dialog UI."""
        layout = QVBoxLayout(self)
        
        # Settings tabs
        self.tab_widget = QTabWidget()
        
        # General settings
        self.setup_general_tab()
        
        # Performance settings
        self.setup_performance_tab()
        
        # Export settings
        self.setup_export_tab()
        
        layout.addWidget(self.tab_widget)
        
        # Dialog buttons
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel | QDialogButtonBox.RestoreDefaults
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        button_box.button(QDialogButtonBox.RestoreDefaults).clicked.connect(self.restore_defaults)
        
        layout.addWidget(button_box)
    
    def setup_general_tab(self):
        """Setup general settings tab."""
        general_tab = QWidget()
        layout = QVBoxLayout(general_tab)
        
        # Application settings
        app_group = QGroupBox("Application")
        app_layout = QFormLayout(app_group)
        
        self.auto_save = QCheckBox("Auto-save results")
        app_layout.addRow("Auto-save:", self.auto_save)
        
        self.confirm_exit = QCheckBox("Confirm before exit")
        app_layout.addRow("Confirm exit:", self.confirm_exit)
        
        self.check_updates = QCheckBox("Check for updates")
        app_layout.addRow("Updates:", self.check_updates)
        
        layout.addWidget(app_group)
        
        # Display settings
        display_group = QGroupBox("Display")
        display_layout = QFormLayout(display_group)
        
        self.theme = QComboBox()
        self.theme.addItems(["System", "Light", "Dark"])
        display_layout.addRow("Theme:", self.theme)
        
        self.font_size = QSpinBox()
        self.font_size.setRange(8, 16)
        self.font_size.setSuffix(" pt")
        display_layout.addRow("Font Size:", self.font_size)
        
        layout.addWidget(display_group)
        
        layout.addStretch()
        self.tab_widget.addTab(general_tab, "General")
    
    def setup_performance_tab(self):
        """Setup performance settings tab."""
        performance_tab = QWidget()
        layout = QVBoxLayout(performance_tab)
        
        # Threading settings
        thread_group = QGroupBox("Threading")
        thread_layout = QFormLayout(thread_group)
        
        self.max_threads = QSpinBox()
        self.max_threads.setRange(1, 16)
        thread_layout.addRow("Max Threads:", self.max_threads)
        
        self.use_multiprocessing = QCheckBox("Use multiprocessing")
        thread_layout.addRow("Multiprocessing:", self.use_multiprocessing)
        
        layout.addWidget(thread_group)
        
        # Memory settings
        memory_group = QGroupBox("Memory")
        memory_layout = QFormLayout(memory_group)
        
        self.cache_size = QSpinBox()
        self.cache_size.setRange(100, 2000)
        self.cache_size.setSuffix(" MB")
        memory_layout.addRow("Cache Size:", self.cache_size)
        
        self.clear_cache_on_exit = QCheckBox("Clear cache on exit")
        memory_layout.addRow("Clear cache:", self.clear_cache_on_exit)
        
        layout.addWidget(memory_group)
        
        layout.addStretch()
        self.tab_widget.addTab(performance_tab, "Performance")
    
    def setup_export_tab(self):
        """Setup export settings tab."""
        export_tab = QWidget()
        layout = QVBoxLayout(export_tab)
        
        # Default export settings
        export_group = QGroupBox("Default Export Settings")
        export_layout = QFormLayout(export_group)
        
        self.default_format = QComboBox()
        self.default_format.addItems(["CSV", "Excel", "PDF", "JSON"])
        export_layout.addRow("Default Format:", self.default_format)
        
        self.include_metadata = QCheckBox("Include metadata")
        export_layout.addRow("Metadata:", self.include_metadata)
        
        self.include_warnings = QCheckBox("Include warnings")
        export_layout.addRow("Warnings:", self.include_warnings)
        
        layout.addWidget(export_group)
        
        # File naming
        naming_group = QGroupBox("File Naming")
        naming_layout = QFormLayout(naming_group)
        
        self.filename_template = QLineEdit()
        self.filename_template.setPlaceholderText("primers_{date}_{time}")
        naming_layout.addRow("Template:", self.filename_template)
        
        self.auto_timestamp = QCheckBox("Auto-add timestamp")
        naming_layout.addRow("Timestamp:", self.auto_timestamp)
        
        layout.addWidget(naming_group)
        
        layout.addStretch()
        self.tab_widget.addTab(export_tab, "Export")
    
    def load_current_settings(self):
        """Load current application settings."""
        # Set default values
        self.auto_save.setChecked(True)
        self.confirm_exit.setChecked(True)
        self.check_updates.setChecked(True)
        self.theme.setCurrentText("System")
        self.font_size.setValue(10)
        self.max_threads.setValue(4)
        self.use_multiprocessing.setChecked(False)
        self.cache_size.setValue(500)
        self.clear_cache_on_exit.setChecked(True)
        self.default_format.setCurrentText("CSV")
        self.include_metadata.setChecked(True)
        self.include_warnings.setChecked(True)
        self.filename_template.setText("primers_{date}_{time}")
        self.auto_timestamp.setChecked(True)
    
    def restore_defaults(self):
        """Restore default settings."""
        self.load_current_settings()
    
    def get_settings(self) -> Dict[str, Any]:
        """Get current settings as dictionary."""
        return {
            'auto_save': self.auto_save.isChecked(),
            'confirm_exit': self.confirm_exit.isChecked(),
            'check_updates': self.check_updates.isChecked(),
            'theme': self.theme.currentText(),
            'font_size': self.font_size.value(),
            'max_threads': self.max_threads.value(),
            'use_multiprocessing': self.use_multiprocessing.isChecked(),
            'cache_size': self.cache_size.value(),
            'clear_cache_on_exit': self.clear_cache_on_exit.isChecked(),
            'default_format': self.default_format.currentText(),
            'include_metadata': self.include_metadata.isChecked(),
            'include_warnings': self.include_warnings.isChecked(),
            'filename_template': self.filename_template.text(),
            'auto_timestamp': self.auto_timestamp.isChecked(),
        }


class ExportWorker(QThread):
    """Worker thread for exporting results."""
    
    progress_updated = Signal(int, str)
    export_completed = Signal(str)
    export_failed = Signal(str)
    
    def __init__(self, primer_sets, file_path, format_type, options):
        super().__init__()
        self.primer_sets = primer_sets
        self.file_path = file_path
        self.format_type = format_type
        self.options = options
        self.logger = get_logger(__name__)
    
    def run(self):
        """Run export in background thread."""
        try:
            self.progress_updated.emit(10, "Preparing export...")
            
            if self.format_type == "CSV":
                self.export_csv()
            elif self.format_type == "Excel":
                self.export_excel()
            elif self.format_type == "PDF":
                self.export_pdf()
            elif self.format_type == "JSON":
                self.export_json()
            
            self.progress_updated.emit(100, "Export completed!")
            self.export_completed.emit(self.file_path)
            
        except Exception as e:
            self.logger.error(f"Export failed: {e}")
            self.export_failed.emit(str(e))
    
    def export_csv(self):
        """Export to CSV format."""
        import csv
        
        self.progress_updated.emit(30, "Writing CSV data...")
        
        with open(self.file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            header = [
                "Set", "Overall_Score", "Tm_Uniformity", "Specificity_Score",
                "F3_Sequence", "F3_Tm", "F3_GC", "F3_Position",
                "B3_Sequence", "B3_Tm", "B3_GC", "B3_Position",
                "FIP_Sequence", "FIP_Tm", "FIP_GC", "FIP_Position",
                "BIP_Sequence", "BIP_Tm", "BIP_GC", "BIP_Position",
                "Amplicon_Size"
            ]
            
            if self.options.get('include_warnings', True):
                header.append("Warnings")
            
            writer.writerow(header)
            
            # Write data
            for i, primer_set in enumerate(self.primer_sets):
                self.progress_updated.emit(30 + (i * 60 // len(self.primer_sets)), 
                                         f"Processing set {i+1}...")
                
                row = [
                    i + 1,
                    f"{primer_set.overall_score:.3f}",
                    f"{primer_set.tm_uniformity:.1f}",
                    f"{primer_set.specificity_score:.3f}",
                    primer_set.f3.sequence,
                    f"{primer_set.f3.tm:.1f}",
                    f"{primer_set.f3.gc_content:.1f}",
                    f"{primer_set.f3.start_pos}-{primer_set.f3.end_pos}",
                    primer_set.b3.sequence,
                    f"{primer_set.b3.tm:.1f}",
                    f"{primer_set.b3.gc_content:.1f}",
                    f"{primer_set.b3.start_pos}-{primer_set.b3.end_pos}",
                    primer_set.fip.sequence,
                    f"{primer_set.fip.tm:.1f}",
                    f"{primer_set.fip.gc_content:.1f}",
                    f"{primer_set.fip.start_pos}-{primer_set.fip.end_pos}",
                    primer_set.bip.sequence,
                    f"{primer_set.bip.tm:.1f}",
                    f"{primer_set.bip.gc_content:.1f}",
                    f"{primer_set.bip.start_pos}-{primer_set.bip.end_pos}",
                    primer_set.f2_b2_amplicon_size
                ]
                
                if self.options.get('include_warnings', True):
                    row.append("; ".join(primer_set.warnings))
                
                writer.writerow(row)
    
    def export_excel(self):
        """Export to Excel format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")
        
        self.progress_updated.emit(30, "Creating Excel workbook...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Primer Sets"
        
        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write headers
        headers = [
            "Set", "Overall Score", "Tm Uniformity", "Specificity Score",
            "F3 Sequence", "F3 Tm", "F3 GC%", "F3 Position",
            "B3 Sequence", "B3 Tm", "B3 GC%", "B3 Position",
            "FIP Sequence", "FIP Tm", "FIP GC%", "FIP Position",
            "BIP Sequence", "BIP Tm", "BIP GC%", "BIP Position",
            "Amplicon Size", "Warnings"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for i, primer_set in enumerate(self.primer_sets, 2):
            self.progress_updated.emit(30 + ((i-2) * 60 // len(self.primer_sets)), 
                                     f"Processing set {i-1}...")
            
            ws.cell(row=i, column=1, value=i-1)
            ws.cell(row=i, column=2, value=primer_set.overall_score)
            ws.cell(row=i, column=3, value=primer_set.tm_uniformity)
            ws.cell(row=i, column=4, value=primer_set.specificity_score)
            
            # F3 data
            ws.cell(row=i, column=5, value=primer_set.f3.sequence)
            ws.cell(row=i, column=6, value=primer_set.f3.tm)
            ws.cell(row=i, column=7, value=primer_set.f3.gc_content)
            ws.cell(row=i, column=8, value=f"{primer_set.f3.start_pos}-{primer_set.f3.end_pos}")
            
            # B3 data
            ws.cell(row=i, column=9, value=primer_set.b3.sequence)
            ws.cell(row=i, column=10, value=primer_set.b3.tm)
            ws.cell(row=i, column=11, value=primer_set.b3.gc_content)
            ws.cell(row=i, column=12, value=f"{primer_set.b3.start_pos}-{primer_set.b3.end_pos}")
            
            # FIP data
            ws.cell(row=i, column=13, value=primer_set.fip.sequence)
            ws.cell(row=i, column=14, value=primer_set.fip.tm)
            ws.cell(row=i, column=15, value=primer_set.fip.gc_content)
            ws.cell(row=i, column=16, value=f"{primer_set.fip.start_pos}-{primer_set.fip.end_pos}")
            
            # BIP data
            ws.cell(row=i, column=17, value=primer_set.bip.sequence)
            ws.cell(row=i, column=18, value=primer_set.bip.tm)
            ws.cell(row=i, column=19, value=primer_set.bip.gc_content)
            ws.cell(row=i, column=20, value=f"{primer_set.bip.start_pos}-{primer_set.bip.end_pos}")
            
            ws.cell(row=i, column=21, value=primer_set.f2_b2_amplicon_size)
            ws.cell(row=i, column=22, value="; ".join(primer_set.warnings))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.file_path)
    
    def export_pdf(self):
        """Export to PDF format."""
        # This would require a PDF library like reportlab
        # For now, just create a text-based PDF
        raise NotImplementedError("PDF export not yet implemented")
    
    def export_json(self):
        """Export to JSON format."""
        import json
        
        self.progress_updated.emit(30, "Converting to JSON...")
        
        data = {
            "primer_sets": [],
            "metadata": {
                "total_sets": len(self.primer_sets),
                "export_format": "JSON",
                "timestamp": __import__('datetime').datetime.now().isoformat()
            }
        }
        
        for i, primer_set in enumerate(self.primer_sets):
            self.progress_updated.emit(30 + (i * 60 // len(self.primer_sets)), 
                                     f"Processing set {i+1}...")
            
            set_data = {
                "set_number": i + 1,
                "overall_score": primer_set.overall_score,
                "tm_uniformity": primer_set.tm_uniformity,
                "specificity_score": primer_set.specificity_score,
                "amplicon_size": primer_set.f2_b2_amplicon_size,
                "primers": {
                    "F3": {
                        "sequence": primer_set.f3.sequence,
                        "tm": primer_set.f3.tm,
                        "gc_content": primer_set.f3.gc_content,
                        "position": f"{primer_set.f3.start_pos}-{primer_set.f3.end_pos}"
                    },
                    "B3": {
                        "sequence": primer_set.b3.sequence,
                        "tm": primer_set.b3.tm,
                        "gc_content": primer_set.b3.gc_content,
                        "position": f"{primer_set.b3.start_pos}-{primer_set.b3.end_pos}"
                    },
                    "FIP": {
                        "sequence": primer_set.fip.sequence,
                        "tm": primer_set.fip.tm,
                        "gc_content": primer_set.fip.gc_content,
                        "position": f"{primer_set.fip.start_pos}-{primer_set.fip.end_pos}"
                    },
                    "BIP": {
                        "sequence": primer_set.bip.sequence,
                        "tm": primer_set.bip.tm,
                        "gc_content": primer_set.bip.gc_content,
                        "position": f"{primer_set.bip.start_pos}-{primer_set.bip.end_pos}"
                    }
                },
                "warnings": primer_set.warnings
            }
            
            data["primer_sets"].append(set_data)
        
        with open(self.file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)


class ExportDialog(QDialog):
    """Dialog for exporting primer design results."""
    
    def __init__(self, primer_sets: List[LampPrimerSet], parent=None):
        super().__init__(parent)
        self.primer_sets = primer_sets
        self.logger = get_logger(__name__)
        
        self.setWindowTitle("Export Results")
        self.setMinimumSize(500, 400)
        
        self.setup_ui()
        self.connect_signals()
    
    def setup_ui(self):
        """Setup the export dialog UI."""
        layout = QVBoxLayout(self)
        
        # Export format selection
        format_group = QGroupBox("Export Format")
        format_layout = QVBoxLayout(format_group)
        
        self.format_group = QButtonGroup()
        
        self.csv_radio = QRadioButton("CSV (Comma Separated Values)")
        self.csv_radio.setChecked(True)
        self.format_group.addButton(self.csv_radio, 0)
        format_layout.addWidget(self.csv_radio)
        
        self.excel_radio = QRadioButton("Excel Spreadsheet (.xlsx)")
        self.format_group.addButton(self.excel_radio, 1)
        format_layout.addWidget(self.excel_radio)
        
        self.json_radio = QRadioButton("JSON (JavaScript Object Notation)")
        self.format_group.addButton(self.json_radio, 2)
        format_layout.addWidget(self.json_radio)
        
        self.pdf_radio = QRadioButton("PDF Report (Not yet available)")
        self.pdf_radio.setEnabled(False)
        self.format_group.addButton(self.pdf_radio, 3)
        format_layout.addWidget(self.pdf_radio)
        
        layout.addWidget(format_group)
        
        # Export options
        options_group = QGroupBox("Export Options")
        options_layout = QVBoxLayout(options_group)
        
        self.include_metadata = QCheckBox("Include metadata")
        self.include_metadata.setChecked(True)
        options_layout.addWidget(self.include_metadata)
        
        self.include_warnings = QCheckBox("Include warnings")
        self.include_warnings.setChecked(True)
        options_layout.addWidget(self.include_warnings)
        
        self.include_all_sets = QCheckBox("Include all primer sets")
        self.include_all_sets.setChecked(True)
        options_layout.addWidget(self.include_all_sets)
        
        layout.addWidget(options_group)
        
        # File selection
        file_group = QGroupBox("Output File")
        file_layout = QHBoxLayout(file_group)
        
        self.file_path = QLineEdit()
        self.file_path.setPlaceholderText("Select output file...")
        file_layout.addWidget(self.file_path)
        
        self.browse_button = QPushButton("Browse...")
        self.browse_button.clicked.connect(self.browse_file)
        file_layout.addWidget(self.browse_button)
        
        layout.addWidget(file_group)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("")
        self.status_label.setVisible(False)
        layout.addWidget(self.status_label)
        
        # Dialog buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.export_button = QPushButton("Export")
        self.export_button.clicked.connect(self.start_export)
        button_layout.addWidget(self.export_button)
        
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_button)
        
        layout.addLayout(button_layout)
    
    def connect_signals(self):
        """Connect dialog signals."""
        self.format_group.buttonClicked.connect(self.on_format_changed)
    
    def on_format_changed(self):
        """Handle format selection changes."""
        # Update file extension suggestion
        current_path = self.file_path.text()
        if current_path:
            base_path = Path(current_path).with_suffix('')
            
            if self.csv_radio.isChecked():
                self.file_path.setText(str(base_path.with_suffix('.csv')))
            elif self.excel_radio.isChecked():
                self.file_path.setText(str(base_path.with_suffix('.xlsx')))
            elif self.json_radio.isChecked():
                self.file_path.setText(str(base_path.with_suffix('.json')))
    
    def browse_file(self):
        """Open file browser for output file selection."""
        if self.csv_radio.isChecked():
            file_filter = "CSV files (*.csv);;All files (*.*)"
            default_ext = ".csv"
        elif self.excel_radio.isChecked():
            file_filter = "Excel files (*.xlsx);;All files (*.*)"
            default_ext = ".xlsx"
        elif self.json_radio.isChecked():
            file_filter = "JSON files (*.json);;All files (*.*)"
            default_ext = ".json"
        else:
            file_filter = "All files (*.*)"
            default_ext = ""
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Export File",
            f"primer_results{default_ext}",
            file_filter
        )
        
        if file_path:
            self.file_path.setText(file_path)
    
    def start_export(self):
        """Start the export process."""
        file_path = self.file_path.text().strip()
        if not file_path:
            QMessageBox.warning(self, "Warning", "Please select an output file.")
            return
        
        # Determine format
        if self.csv_radio.isChecked():
            format_type = "CSV"
        elif self.excel_radio.isChecked():
            format_type = "Excel"
        elif self.json_radio.isChecked():
            format_type = "JSON"
        else:
            QMessageBox.warning(self, "Warning", "Please select an export format.")
            return
        
        # Get options
        options = {
            'include_metadata': self.include_metadata.isChecked(),
            'include_warnings': self.include_warnings.isChecked(),
            'include_all_sets': self.include_all_sets.isChecked(),
        }
        
        # Setup UI for export
        self.export_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.status_label.setVisible(True)
        self.progress_bar.setValue(0)
        
        # Start export worker
        self.export_worker = ExportWorker(self.primer_sets, file_path, format_type, options)
        self.export_worker.progress_updated.connect(self.on_export_progress)
        self.export_worker.export_completed.connect(self.on_export_completed)
        self.export_worker.export_failed.connect(self.on_export_failed)
        self.export_worker.start()
    
    def on_export_progress(self, percentage, message):
        """Handle export progress updates."""
        self.progress_bar.setValue(percentage)
        self.status_label.setText(message)
    
    def on_export_completed(self, file_path):
        """Handle successful export completion."""
        self.progress_bar.setVisible(False)
        self.status_label.setVisible(False)
        self.export_button.setEnabled(True)
        
        QMessageBox.information(
            self,
            "Export Successful",
            f"Results exported successfully to:\n{file_path}"
        )
        
        self.accept()
    
    def on_export_failed(self, error_message):
        """Handle export failure."""
        self.progress_bar.setVisible(False)
        self.status_label.setVisible(False)
        self.export_button.setEnabled(True)
        
        QMessageBox.critical(
            self,
            "Export Failed",
            f"Export failed:\n\n{error_message}"
        )
